"""Export endpoints — JSON, CSV, and Excel (.xlsx) session reports."""

import csv
import io
import json
from datetime import datetime, timezone

from flask import Blueprint, jsonify, make_response, request

from models import Document, DocumentEntity, DocumentSummary, DocumentText, Session

export_bp = Blueprint("export", __name__)


def _session_data(session_id: str) -> dict | None:
    session = Session.query.get(session_id)
    if not session:
        return None

    docs = Document.query.filter_by(session_id=session_id).order_by(Document.uploaded_at).all()

    documents = []
    for doc in docs:
        dt = DocumentText.query.filter_by(document_id=doc.id).first()
        summary = DocumentSummary.query.filter_by(document_id=doc.id).first()
        entities = (
            DocumentEntity.query
            .filter_by(document_id=doc.id)
            .order_by(DocumentEntity.entity_type, DocumentEntity.label)
            .all()
        )
        documents.append({
            "id": doc.id,
            "filename": doc.filename,
            "filetype": doc.filetype,
            "status": doc.status,
            "uploaded_at": doc.uploaded_at.isoformat(),
            "word_count": dt.word_count if dt else None,
            "page_count": dt.page_count if dt else None,
            "chunk_count": len(doc.chunks),
            "headline": summary.headline if summary else "",
            "summary_text": summary.summary_text if summary else "",
            "key_points": summary.key_points if summary else [],
            "entities": [e.to_dict() for e in entities],
        })

    return {
        "session": session.to_dict(),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "document_count": len(documents),
        "documents": documents,
    }


@export_bp.route("/sessions/<session_id>/export/json", methods=["GET"])
def export_json(session_id: str):
    data = _session_data(session_id)
    if data is None:
        return jsonify({"error": "Session not found"}), 404

    payload = json.dumps(data, indent=2, ensure_ascii=False)
    resp = make_response(payload)
    resp.headers["Content-Type"] = "application/json"
    safe_name = data["session"]["name"].replace(" ", "_")[:40]
    resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}_export.json"'
    return resp


@export_bp.route("/sessions/<session_id>/export/csv", methods=["GET"])
def export_csv(session_id: str):
    data = _session_data(session_id)
    if data is None:
        return jsonify({"error": "Session not found"}), 404

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Choose sheet via ?sheet= param (entities | summaries | documents)
    sheet = request.args.get("sheet", "entities")

    if sheet == "summaries":
        writer.writerow(["Document", "Headline", "Summary", "Key Points"])
        for doc in data["documents"]:
            writer.writerow([
                doc["filename"],
                doc["headline"],
                doc["summary_text"],
                " | ".join(doc["key_points"]),
            ])
    elif sheet == "documents":
        writer.writerow(["Filename", "Type", "Status", "Pages", "Words", "Chunks", "Uploaded"])
        for doc in data["documents"]:
            writer.writerow([
                doc["filename"],
                doc["filetype"].upper(),
                doc["status"],
                doc["page_count"] or "",
                doc["word_count"] or "",
                doc["chunk_count"],
                doc["uploaded_at"][:10],
            ])
    else:  # entities (default)
        writer.writerow(["Document", "Entity Type", "Label", "Value"])
        for doc in data["documents"]:
            for ent in doc["entities"]:
                writer.writerow([
                    doc["filename"],
                    ent["entity_type"],
                    ent["label"],
                    ent["value"],
                ])

    payload = buf.getvalue()
    resp = make_response(payload)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    safe_name = data["session"]["name"].replace(" ", "_")[:40]
    resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}_{sheet}.csv"'
    return resp


@export_bp.route("/sessions/<session_id>/export/xlsx", methods=["GET"])
def export_xlsx(session_id: str):
    data = _session_data(session_id)
    if data is None:
        return jsonify({"error": "Session not found"}), 404

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    _sheet_overview(wb, data)
    _sheet_summaries(wb, data)
    _sheet_entities(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = data["session"]["name"].replace(" ", "_")[:40]
    resp.headers["Content-Disposition"] = f'attachment; filename="{safe_name}_report.xlsx"'
    return resp


# ── Excel sheet builders ──────────────────────────────────────────────────────

def _hdr(ws, row: int, values: list[str], bg: str = "1E3A5F"):
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _row(ws, row: int, values: list, wrap: bool = False):
    from openpyxl.styles import Alignment
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def _sheet_overview(wb, data: dict):
    ws = wb.active
    ws.title = "Documents"
    ws.row_dimensions[1].height = 20

    _hdr(ws, 1, ["Filename", "Type", "Status", "Pages", "Words", "Chunks", "Uploaded"])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    for i, doc in enumerate(data["documents"], 2):
        _row(ws, i, [
            doc["filename"],
            doc["filetype"].upper(),
            doc["status"].capitalize(),
            doc["page_count"],
            doc["word_count"],
            doc["chunk_count"],
            doc["uploaded_at"][:10],
        ])


def _sheet_summaries(wb, data: dict):
    from openpyxl.styles import Font
    ws = wb.create_sheet("Summaries")
    _hdr(ws, 1, ["Document", "Headline", "Summary", "Key Points"], bg="2D6A4F")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50

    for i, doc in enumerate(data["documents"], 2):
        _row(ws, i, [
            doc["filename"],
            doc["headline"],
            doc["summary_text"],
            "\n".join(f"- {p}" for p in doc["key_points"]),
        ], wrap=True)
        ws.row_dimensions[i].height = max(15 * len(doc["key_points"]), 20)


def _sheet_entities(wb, data: dict):
    ws = wb.create_sheet("Entities")
    _hdr(ws, 1, ["Document", "Entity Type", "Label", "Value"], bg="6B2D8B")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40

    row = 2
    for doc in data["documents"]:
        for ent in doc["entities"]:
            _row(ws, row, [
                doc["filename"],
                ent["entity_type"].replace("_", " ").title(),
                ent["label"],
                ent["value"],
            ])
            row += 1
